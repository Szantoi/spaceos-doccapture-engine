"""A munkafüzet-adapter — a gyorsítótár-olvasás és a futtatás-mentesség mérése.

A LEGFONTOSABB TESZT EBBEN A FÁJLBAN
-----------------------------------
`test_a_kepletet_NEM_szamoljuk_ujra`: egy `=1+1` képlet mellé **szándékosan
hibás** tárolt értéket (99) injektálunk a fájlba. Ha az adapter 99-et ad, akkor
**bizonyítottan a gyorsítótárat olvassa** és nem értékel ki képletet. Ha 2-t
adna, kiértékelt — és akkor a determinizmus (M11) elveszett.

Ez a mutációs bizonyítás megfelelője itt: nem azt mérjük, hogy „működik", hanem
hogy a mérés **meg tudja különböztetni** a két viselkedést.
"""

from __future__ import annotations

import shutil
import tempfile
import unittest
import zipfile
from pathlib import Path

from doccapture.core.config import CaptureConfig
from doccapture.core.errors import SourceUnreadableError
from doccapture.core.models import Confidence
from doccapture.core.tabular import ColumnSpec, ColumnType, TableSchema, TabularOptions
from doccapture.infrastructure.tabular.workbook import WorkbookTabularReader

try:  # pragma: no cover - telepitesi kerdes
    import openpyxl

    HAS_OPENPYXL = True
except ImportError:  # pragma: no cover
    HAS_OPENPYXL = False

SCHEMA = TableSchema(
    columns=(
        ColumnSpec("kod", ("Kód",), ColumnType.TEXT, required=True),
        ColumnSpec("megnevezes", ("Megnevezés",), ColumnType.TEXT),
        ColumnSpec("ertek", ("Érték",), ColumnType.NUMBER),
    ),
    identity_keys=("kod",),
)


def _inject_cached_value(path: Path, cell_ref: str, cached: str) -> None:
    """Tárolt értéket injektál egy képlet-cella mellé, a fájl újraírásával.

    Miért kézzel: az olvasó könyvtár nem tud képletet ÉS tárolt értéket együtt
    kiírni — pedig épp ez az az állapot, amit egy táblázatkezelővel mentett
    fájl mindig tartalmaz. Enélkül nem lehetne megmérni, hogy melyiket olvassuk.
    """
    with zipfile.ZipFile(path) as source:
        entries = {name: source.read(name) for name in source.namelist()}

    sheet_name = next(n for n in entries if n.startswith("xl/worksheets/sheet"))
    xml = entries[sheet_name].decode("utf-8")

    marker = f'<c r="{cell_ref}"'
    start = xml.index(marker)
    close = xml.index("</f>", start) + len("</f>")
    entries[sheet_name] = (xml[:close] + f"<v>{cached}</v>" + xml[close:]).encode("utf-8")

    path.unlink()
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as target:
        for name, data in entries.items():
            target.writestr(name, data)


class _Workbook:
    """Eldobható forrás-mappa egyetlen munkafüzettel."""

    def __init__(self, name: str = "lista.xlsx") -> None:
        self._tmp = tempfile.TemporaryDirectory()
        self.root = Path(self._tmp.name)
        self.name = name
        self.path = self.root / name

    def save(self, workbook) -> None:
        workbook.save(self.path)

    def reader(self, **options) -> WorkbookTabularReader:
        config = CaptureConfig(input_root=str(self.root), tabular=TabularOptions(**options))
        return WorkbookTabularReader(config)

    def config(self, **kwargs) -> CaptureConfig:
        return CaptureConfig(input_root=str(self.root), **kwargs)

    def close(self) -> None:
        self._tmp.cleanup()


def _simple_workbook(rows: list[list[object]], title: str = "Adatok"):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = title
    for row in rows:
        sheet.append(row)
    return workbook


@unittest.skipUnless(HAS_OPENPYXL, "a munkafuzet-ut a `tabular` extrat igenyli")
class ReadingTests(unittest.TestCase):
    def setUp(self) -> None:
        self.fixture = _Workbook()
        self.addCleanup(self.fixture.close)
        self.fixture.save(
            _simple_workbook(
                [
                    ["Kód", "Megnevezés", "Érték"],
                    ["A-1", "Első tétel", 12.5],
                    ["A-2", "Második tétel", 3],
                ]
            )
        )

    def test_a_sorok_a_belso_kulcsokkal_es_natív_tipussal_jonnek(self) -> None:
        result = self.fixture.reader().read(self.fixture.name, SCHEMA)

        self.assertEqual(len(result.rows), 2)
        self.assertEqual(result.rows[0]["kod"].value, "A-1")
        self.assertEqual(result.rows[0]["ertek"].value, 12.5)
        self.assertIs(result.rows[0]["ertek"].confidence, Confidence.CONFIRMED)

    def test_a_bizonyitek_a_LAPOT_is_megnevezi(self) -> None:
        """Egy munkafüzetben több lap van: a sorszám lap nélkül nem hely."""
        result = self.fixture.reader().read(self.fixture.name, SCHEMA)

        evidence = result.rows[0]["kod"].evidence
        self.assertEqual(evidence.locator, "Adatok!R2C1")
        self.assertEqual(evidence.relative_path, "lista.xlsx")
        self.assertTrue(evidence.content_hash.startswith("sha256:"))


@unittest.skipUnless(HAS_OPENPYXL, "a munkafuzet-ut a `tabular` extrat igenyli")
class FormulaTests(unittest.TestCase):
    """A szelet legdrágább leletje: a képlet gyorsítótár nélkül NEM üres cella."""

    def test_a_kepletet_NEM_szamoljuk_ujra(self) -> None:
        """A `=1+1` mellé 99-et injektálunk tárolt értékként.

        Ha 99-et kapunk: a gyorsítótárat olvassuk (helyes, M11).
        Ha 2-t kapnánk: kiértékeltük a képletet, és a determinizmus elveszett.
        A teszt attól bizonyíték, hogy a két viselkedés **megkülönböztethető**.
        """
        fixture = _Workbook()
        self.addCleanup(fixture.close)
        fixture.save(
            _simple_workbook([["Kód", "Megnevezés", "Érték"], ["A-1", "Tétel", "=1+1"]])
        )
        _inject_cached_value(fixture.path, "C2", "99")

        result = fixture.reader().read(fixture.name, SCHEMA)

        self.assertEqual(result.rows[0]["ertek"].value, 99.0)
        self.assertNotEqual(
            result.rows[0]["ertek"].value, 2.0, "a kepletet kiertekeltuk — M11 serult"
        )

    def test_a_gyorsitotar_nelkuli_keplet_HIANY_indokkal(self) -> None:
        """Az olvasó itt `None`-t ad, ami megkülönböztethetetlen az üres cellától.
        A naiv megoldás vagy `"None"`-t írna be, vagy üres sornak vennné — mindkettő
        néma adatvesztés."""
        fixture = _Workbook()
        self.addCleanup(fixture.close)
        fixture.save(
            _simple_workbook([["Kód", "Megnevezés", "Érték"], ["A-1", "Tétel", "=1+1"]])
        )

        result = fixture.reader().read(fixture.name, SCHEMA)
        cell = result.rows[0]["ertek"]

        self.assertIsNone(cell.value)
        self.assertIs(cell.confidence, Confidence.MISSING)
        self.assertIn("gyorsítótár", cell.note)
        self.assertNotEqual(cell.note, "üres cella")

    def test_a_gyorsitotar_nelkuli_keplet_NEM_teszi_ures_sorra_a_sort(self) -> None:
        """Ha az azonosító oszlopban áll ilyen képlet, a sor kiesne — csendben."""
        fixture = _Workbook()
        self.addCleanup(fixture.close)
        fixture.save(
            _simple_workbook([["Kód", "Megnevezés", "Érték"], ['=CONCATENATE("A-",1)', "T", 1]])
        )

        result = fixture.reader().read(fixture.name, SCHEMA)

        # A sor MEGMARAD (nem esik ki), es a hianyt kimondja.
        self.assertEqual(result.skipped_blank_rows, 0)
        self.assertEqual(len(result.rows), 1)
        self.assertIn("gyorsítótár", result.rows[0]["kod"].note)
        self.assertTrue(result.needs_human)

    def test_a_valodi_ures_cella_MAS_indokot_kap(self) -> None:
        """A két eset szétválasztása a lelet lényege — ha egy indokot adnánk,
        az ember nem tudná, hogy a forrást kell-e megnyitni."""
        fixture = _Workbook()
        self.addCleanup(fixture.close)
        fixture.save(
            _simple_workbook([["Kód", "Megnevezés", "Érték"], ["A-1", None, None]])
        )

        result = fixture.reader().read(fixture.name, SCHEMA)
        self.assertEqual(result.rows[0]["ertek"].note, "üres cella")


@unittest.skipUnless(HAS_OPENPYXL, "a munkafuzet-ut a `tabular` extrat igenyli")
class SheetSelectionTests(unittest.TestCase):
    def test_a_lap_nev_szerint_valaszthato(self) -> None:
        fixture = _Workbook()
        self.addCleanup(fixture.close)
        workbook = _simple_workbook([["Kód", "Megnevezés", "Érték"], ["ELSO", "T", 1]], "Egy")
        second = workbook.create_sheet("Ketto")
        for row in [["Kód", "Megnevezés", "Érték"], ["MASODIK", "T", 2]]:
            second.append(row)
        fixture.save(workbook)

        self.assertEqual(
            fixture.reader(sheet_name="Ketto").read(fixture.name, SCHEMA).rows[0]["kod"].value,
            "MASODIK",
        )

    def test_nem_letezo_lap_KIMONDOTT_hiba_a_lapok_felsorolasaval(self) -> None:
        fixture = _Workbook()
        self.addCleanup(fixture.close)
        fixture.save(_simple_workbook([["Kód", "Megnevezés", "Érték"]], "Egy"))

        with self.assertRaises(SourceUnreadableError) as ctx:
            fixture.reader(sheet_name="Nincs").read(fixture.name, SCHEMA)
        self.assertIn("Egy", str(ctx.exception))

    def test_lapnev_nelkul_az_AKTIV_lap_jon(self) -> None:
        """Az aktív lap kényelmes, de MENTÉSKOR változik — ezért nem stabil.
        A doc ezt kimondja; itt csak azt mérjük, hogy a viselkedés az, amit ígér."""
        fixture = _Workbook()
        self.addCleanup(fixture.close)
        fixture.save(_simple_workbook([["Kód", "Megnevezés", "Érték"], ["A-1", "T", 1]], "Egy"))

        self.assertEqual(fixture.reader().read(fixture.name, SCHEMA).rows[0]["kod"].value, "A-1")


@unittest.skipUnless(HAS_OPENPYXL, "a munkafuzet-ut a `tabular` extrat igenyli")
class ActiveContentTests(unittest.TestCase):
    def test_a_futtatas_keres_NEM_nemul_el(self) -> None:
        """A kapcsoló más adapternek is szólhat, ezért nem dobunk hibát — de
        egy csendben figyelmen kívül hagyott kapcsoló rosszabb, mint egy
        elutasított kérés."""
        fixture = _Workbook()
        self.addCleanup(fixture.close)
        fixture.save(_simple_workbook([["Kód", "Megnevezés", "Érték"], ["A-1", "T", 1]]))

        config = fixture.config(run_active_content=True)
        result = WorkbookTabularReader(config).read(fixture.name, SCHEMA)

        self.assertTrue(
            any("SEMMIT nem" in note for note in result.diagnostics), result.diagnostics
        )

    def test_a_makros_kiterjesztes_olvashato(self) -> None:
        """⚠ **Amit ez NEM bizonyít:** hogy egy VALÓDI makrót nem futtatunk le.
        Ez a fájl makró-kiterjesztésű, de nincs benne makró-projekt — előállítani
        nem tudunk olyat. Amit bizonyít: a `.xlsm` út olvasható, és a
        gyorsítótárból dolgozik."""
        fixture = _Workbook("lista.xlsm")
        self.addCleanup(fixture.close)
        fixture.save(_simple_workbook([["Kód", "Megnevezés", "Érték"], ["A-1", "T", 1]]))

        result = fixture.reader().read(fixture.name, SCHEMA)
        self.assertEqual(result.rows[0]["kod"].value, "A-1")


@unittest.skipUnless(HAS_OPENPYXL, "a munkafuzet-ut a `tabular` extrat igenyli")
class ReadOnlySourceTests(unittest.TestCase):
    """M10/M8: az eredetit nem bántjuk. MÉRVE, nem állítva."""

    def test_a_betoltes_semmit_nem_valtoztat_a_forras_mappaban(self) -> None:
        from doccapture.infrastructure.evidence import content_hash

        fixture = _Workbook()
        self.addCleanup(fixture.close)
        fixture.save(_simple_workbook([["Kód", "Megnevezés", "Érték"], ["A-1", "T", 1]]))

        def snapshot() -> dict[str, tuple[int, int, str]]:
            return {
                path.name: (path.stat().st_size, path.stat().st_mtime_ns, content_hash(path))
                for path in sorted(fixture.root.rglob("*"))
                if path.is_file()
            }

        before = snapshot()
        fixture.reader().read(fixture.name, SCHEMA)
        after = snapshot()

        self.assertEqual(before, after)
        self.assertEqual(len(before), 1)

    def test_a_betoltes_utan_a_fajl_MOZGATHATO(self) -> None:
        """Windowson egy nyitva hagyott fájl-leíró zárolja a fájlt — az az
        ügyfél mappájában látható kár, még ha nem is írunk bele. Ezt nem
        feltételezzük: megmozgatjuk a fájlt a betöltés után."""
        fixture = _Workbook()
        self.addCleanup(fixture.close)
        fixture.save(_simple_workbook([["Kód", "Megnevezés", "Érték"], ["A-1", "T", 1]]))

        fixture.reader().read(fixture.name, SCHEMA)

        target = fixture.root / "athelyezve.xlsx"
        shutil.move(str(fixture.path), str(target))  # zarolas eseten OSError
        self.assertTrue(target.is_file())


if __name__ == "__main__":
    unittest.main()

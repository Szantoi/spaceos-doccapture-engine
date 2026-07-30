"""Munkafüzet olvasása — a tárolt gyorsítótárból, futtatás NÉLKÜL (M11).

A LEGDRÁGÁBB LELET EBBEN A SZELETBEN: A KÉPLET GYORSÍTÓTÁR NÉLKÜL
-----------------------------------------------------------------
Gyorsítótár-módban a képlet-cella a **tárolt** értékét adja. Ha a fájlt soha nem
mentette ki táblázatkezelő (gépi generálás, vagy a képletet utólag írták bele),
akkor **nincs tárolt érték**, és az olvasó `None`-t ad — ami
megkülönböztethetetlen az üres cellától.

Vagyis a naiv megoldás vagy `"None"` szöveget ír be, vagy „ez a sor üres"-nek
veszi. Mindkettő **néma adatvesztés**, és a forrás-prototípus pontosan ebbe
futott volna bele.

**A megoldás:** a fájlt kétszer nyitjuk meg — egyszer gyorsítótár-módban,
egyszer képlet-módban. Ha a gyorsítótár üres, de a cellában képlet áll, az
eredmény kimondott **hiány indokkal**, nem üres cella. A képletet **nem
futtatjuk ki**: egy futtatott képlet ma és holnap mást adhat, tehát a
determinizmus is elveszne — és aktív tartalmat egyáltalán nem futtatunk.

Ennek az áráról: a kétszeri megnyitás megduplázza a memóriaigényt. Ezért van
`max_rows` korlát — de a határértéket **nem méréssel** állítottuk be, és ezt ki
is mondjuk.
"""

from __future__ import annotations

from dataclasses import replace
from typing import Any, Optional

from doccapture.core.config import CaptureConfig
from doccapture.core.errors import SourceUnreadableError
from doccapture.core.observability import get_logger, log_step
from doccapture.core.ports import TabularReader
from doccapture.core.tabular.assembly import build_result
from doccapture.core.tabular.result import TabularReadResult
from doccapture.core.tabular.schema import TableSchema
from doccapture.core.tabular.values import UnreadableCell
from doccapture.infrastructure.evidence import evidence_for, resolve_source, with_locator

_log = get_logger("workbook")

# Jelzo-ertek a kepletre, aminek nincs tarolt erteke. A tipus a MAGBAN lakik
# (`UnreadableCell`), mert ez altalanos fogalom: "van itt tartalom, de nem
# olvashato, es az okot elmondom". Sajat, adapter-lokalis tipus eseten a mag nem
# ertene, mit adtunk be, es `str()`-rel csendben szoveggé alakitana.
FORMULA_WITHOUT_CACHE = UnreadableCell(
    "képlet gyorsítótár nélkül — a fájlt soha nem mentette ki táblázatkezelő, "
    "tehát nincs tárolt eredmény. A képletet NEM futtatjuk ki (M11): egy "
    "futtatott képlet ma és holnap mást adhat."
)


class WorkbookTabularReader(TabularReader):
    """`TabularReader` munkafüzetekre (makrós fájlt is olvas, de nem futtat)."""

    ADAPTER_NAME = "workbook"

    def __init__(self, config: Optional[CaptureConfig] = None) -> None:
        self._config = config or CaptureConfig()
        self._config.tabular.validate()

    def read(self, relative_path: str, schema: TableSchema) -> TabularReadResult:
        openpyxl = _require_openpyxl()
        options = self._config.tabular
        path = resolve_source(self._config.input_root, relative_path)

        diagnostics: list[str] = []
        if self._config.run_active_content:
            # Nem dobunk hibat: a kapcsolo mas adapternek is szolhat. De NEM
            # hallgatjuk el, hogy mi nem teljesitjuk -- egy csendben figyelmen
            # kivul hagyott kapcsolo rosszabb, mint egy elutasitott keres.
            diagnostics.append(
                "aktív tartalom futtatása kérve, de ez az adapter SEMMIT nem "
                "futtat (makró, képlet, lekérdezés, külső hivatkozás) — a tárolt "
                "gyorsítótárat olvassuk"
            )

        try:
            cached = openpyxl.load_workbook(
                path, data_only=True, read_only=True, keep_links=False
            )
            formulas = openpyxl.load_workbook(
                path, data_only=False, read_only=True, keep_links=False
            )
        except Exception as exc:  # az olvaso sokfele hibat dob; mind vegleges
            raise SourceUnreadableError(
                f"A munkafüzet nem olvasható: {relative_path!r} ({exc})"
            ) from exc

        try:
            cached_sheet = self._select_sheet(cached, relative_path)
            formula_sheet = self._select_sheet(formulas, relative_path)
            sheet_title = cached_sheet.title

            grid = self._merge_grids(cached_sheet, formula_sheet)
        finally:
            # read_only modban a fajl-leiro nyitva marad, amig le nem zarjuk.
            # A forras csak olvashato (M10), de egy nyitva hagyott leiro
            # Windowson zarolja a fajlt -- az ugyfel mappajaban ez lathato kar.
            cached.close()
            formulas.close()

        header_index = options.header_row - 1
        if header_index >= len(grid):
            raise SourceUnreadableError(
                f"A fejléc a(z) {options.header_row}. sorban lenne, de a(z) "
                f"{sheet_title!r} lapnak csak {len(grid)} sora van."
            )

        first_data_index = header_index + 1 + options.data_starts_after_header
        data_rows = [
            (index + 1, row) for index, row in enumerate(grid) if index >= first_data_index
        ]

        file_evidence = evidence_for(self._config.input_root, relative_path)
        # `source` RELATIV ut -- abszolutot a naplo-kapu elbuktat.
        log_step(
            _log,
            "workbook.read",
            source=relative_path,
            data_rows_offered=len(data_rows),
            sheet=sheet_title, formula_without_cache=sum(1 for row in grid for cell in row if isinstance(cell, UnreadableCell)),
        )

        result = build_result(
            header_cells=list(grid[header_index]),
            data_rows=data_rows,
            schema=schema,
            options=options,
            file_evidence=file_evidence,
            locator_for=lambda row, column: f"{sheet_title}!R{row}C{column + 1}",
            header_locator=f"{sheet_title}!R{options.header_row}",
            human_only_field_types=self._config.human_only_field_types,
            evidence_with_locator=with_locator,
        )
        # A sajat diagnosztikankat az OSSZEALLITO ele tesszuk, mert az adapter
        # szintu eszrevetel (pl. futtatas-keres) a fajl egeszere all.
        return _with_extra_diagnostics(result, diagnostics)

    # ------------------------------------------------------------------

    def _select_sheet(self, workbook: Any, relative_path: str) -> Any:
        name = self._config.tabular.sheet_name
        if not name:
            return workbook.active
        if name not in workbook.sheetnames:
            raise SourceUnreadableError(
                f"A(z) {name!r} lap nincs a munkafüzetben ({relative_path!r}). "
                f"Elérhető lapok: {list(workbook.sheetnames)}"
            )
        return workbook[name]

    @staticmethod
    def _merge_grids(cached_sheet: Any, formula_sheet: Any) -> list[list[Any]]:
        """A két olvasás összefésülése: ahol a gyorsítótár üres, de képlet van.

        A `zip` szándékosan `strict=False`: a két megnyitás ugyanazt a fájlt
        látja, tehát elvileg azonos alakú — de ha valamiért nem, akkor **inkább
        a rövidebbig** dolgozunk, mint hogy kivételt dobjunk egy olyan
        eltérésre, amit nem is mi okoztunk. A hiányzó rész üres cellának
        látszik, ami `MISSING` — nem téves érték.
        """
        merged: list[list[Any]] = []
        for cached_row, formula_row in zip(
            cached_sheet.iter_rows(values_only=True),
            formula_sheet.iter_rows(values_only=True),
            strict=False,
        ):
            row: list[Any] = []
            for index, value in enumerate(cached_row):
                if value is None and _is_formula(
                    formula_row[index] if index < len(formula_row) else None
                ):
                    row.append(FORMULA_WITHOUT_CACHE)
                else:
                    row.append(value)
            merged.append(row)
        return merged


def _is_formula(value: Any) -> bool:
    """Képlet-e a cella tartalma képlet-módban.

    A szöveges alak `=`-lel kezdődik. A tömb-képletek külön objektumként
    jönnek, ezért a típus nevét is megnézzük — ha csak a sztringre néznénk, a
    tömb-képletek **csendben** üres cellának látszanának.
    """
    if isinstance(value, str):
        return value.startswith("=")
    return type(value).__name__ in {"ArrayFormula", "DataTableFormula"}


def _with_extra_diagnostics(
    result: TabularReadResult, extra: list[str]
) -> TabularReadResult:
    if not extra:
        return result
    return replace(result, diagnostics=[*extra, *result.diagnostics])


def _require_openpyxl() -> Any:
    """A munkafüzet-olvasó külső csomagot igényel — ezt kimondva kérjük.

    Miért nem a modul fejlécében importáljuk: az egész motort telepíthetetlenné
    tenné annak, akinek csak az elválasztott szöveges út kell. A CSV-ág
    **függőség nélküli**, és ez piaci előny — ne dobjuk el egy import miatt.
    """
    try:
        import openpyxl  # noqa: PLC0415 - szandekosan kesleltetett import
    except ImportError as exc:  # pragma: no cover - telepitesi kerdes
        raise SourceUnreadableError(
            "A munkafüzet-olvasáshoz a `tabular` extra kell: "
            "`pip install doccapture-engine[tabular]`. Az elválasztott szöveges "
            "út ettől függetlenül működik, függőség nélkül."
        ) from exc
    return openpyxl
